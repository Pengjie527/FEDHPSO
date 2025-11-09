"""
Step 2 Main Script: 联邦学习优化方法
包含联邦平均、联邦聚类优化等高级聚合方法
"""

import os
import time
from datetime import datetime
from typing import Dict, List
from local_clustering import run_local_clustering_all_sites
from baseline_aggregation import evaluate_global_clustering
from advanced_aggregation import AdvancedCentralServer
from create_visualization import save_visualizations
import matplotlib as plt
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
import numpy as np
plt.rcParams['font.sans-serif'] = ['SimHei']
# 修复 numpy 1.24+ 删除别名导致的兼容问题
if not hasattr(np, 'float'):
    np.float = float
if not hasattr(np, 'int'):
    np.int = int
if not hasattr(np, 'bool'):
    np.bool = bool
if not hasattr(np, 'complex'):
    np.complex = complex
if not hasattr(np, 'object'):
    np.object = object
if not hasattr(np, 'str'):
    np.str = str

def create_output_directory(base_dir: str = "results") -> str:
    """
    创建输出目录
    
    Parameters:
    -----------
    base_dir : str
        基础目录名
        
    Returns:
    --------
    str : 输出目录路径
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join(base_dir, f"step2_optimized_{timestamp}")
    os.makedirs(output_dir, exist_ok=True)
    logger.info(f"输出目录创建: {output_dir}")
    return output_dir


def run_step2_complete_pipeline(
    data_paths: Dict[int, str],
    n_clusters: int = 750,
    random_state: int = 42,
    output_base_dir: str = "results",
    use_first_bloc_only: bool = True
):
    """
    运行完整的 Step 2 流程
    
    Parameters:
    -----------
    data_paths : Dict[int, str]
        站点ID到数据路径的映射
    n_clusters : int
        聚类数量
    random_state : int
        随机种子
    output_base_dir : str
        输出基础目录
    use_first_bloc_only : bool, default=True
        是否只使用每个患者的第一条bloc数据
    """
    logger.info("\n" + "="*80)
    logger.info("开始运行 Step 2: 优化联邦聚类流程")
    if use_first_bloc_only:
        logger.info("数据选择策略: 每个患者只使用第一条bloc数据")
    logger.info("="*80 + "\n")
    
    # 创建输出目录
    output_dir = create_output_directory(output_base_dir)
    
    execution_time = {}
    
    # Step 2.1: 本地聚类
    logger.info("\n### Step 2.1: 各站点本地聚类 ###\n")
    start_time = time.time()
    sites, upload_infos = run_local_clustering_all_sites(
        data_paths=data_paths,
        n_clusters=n_clusters,
        random_state=random_state,
        use_first_bloc_only=use_first_bloc_only
    )
    execution_time['本地聚类'] = time.time() - start_time
    
    # 保存本地聚类结果（不再单独保存每个站点文件）
    # for site in sites:
    #     local_centers_path = os.path.join(output_dir, f'site{site.site_id}_local_centers.npy')
    #     np.save(local_centers_path, site.cluster_centers)
    #     logger.info(f"站点 {site.site_id} 本地聚类中心已保存: {local_centers_path}")
    
    # Step 2.2: 信息上传到中心服务器
    logger.info("\n### Step 2.2: 信息上传 ###\n")
    server = AdvancedCentralServer(n_clusters=n_clusters)
    server.receive_uploads(upload_infos)
    logger.info(f"收到 {len(upload_infos)} 个站点的上传信息")
    
    # Step 2.3: 高级聚合方法
    logger.info("\n### Step 2.3: 高级聚合方法 ###\n")
    
    # 方法1: 匈牙利匹配聚合
    logger.info("执行匈牙利匹配聚合...")
    start_time = time.time()
    global_centers_1 = server.aggregate_with_hungarian_matching()
    execution_time['匈牙利匹配聚合'] = time.time() - start_time
    
    # 方法2: 层次聚类聚合
    logger.info("执行层次聚类聚合...")
    start_time = time.time()
    global_centers_2 = server.aggregate_hierarchical_clustering()
    execution_time['层次聚类聚合'] = time.time() - start_time
    
    # 方法3: 联邦迭代聚合
    logger.info("执行联邦迭代聚合...")
    start_time = time.time()
    global_centers_3 = server.aggregate_federated_iterative(sites, max_iter=20)
    execution_time['联邦迭代聚合'] = time.time() - start_time
    
    # 评估所有方法
    logger.info("评估聚合结果...")
    all_results = {}
    
    start_time = time.time()
    results_1 = evaluate_global_clustering(sites, global_centers_1)
    execution_time['匈牙利匹配评估'] = time.time() - start_time
    all_results['匈牙利匹配聚合'] = results_1
    
    start_time = time.time()
    results_2 = evaluate_global_clustering(sites, global_centers_2)
    execution_time['层次聚类评估'] = time.time() - start_time
    all_results['层次聚类聚合'] = results_2
    
    start_time = time.time()
    results_3 = evaluate_global_clustering(sites, global_centers_3)
    execution_time['联邦迭代评估'] = time.time() - start_time
    all_results['联邦迭代聚合'] = results_3
    
    # 保存最佳结果（基于平均惯性比率，越小越好）
    best_method = min(all_results.keys(), key=lambda k: all_results[k]['inertia_ratio'].mean())
    best_centers = global_centers_1 if best_method == '匈牙利匹配聚合' else (global_centers_2 if best_method == '层次聚类聚合' else global_centers_3)
    
    # 保存最佳结果（npy文件保留，但评估结果整合到Excel）
    np.save(os.path.join(output_dir, 'best_global_centers.npy'), best_centers)
    
    logger.info(f"最佳聚合方法: {best_method}")
    logger.info(f"最佳平均惯性比率: {all_results[best_method]['inertia_ratio'].mean():.4f}")
    
    # 整合所有结果到一个Excel文件
    logger.info("\n### 整合所有结果到Excel文件 ###\n")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "方法对比"
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # 创建汇总表
    row = 1
    ws.cell(row, 1, "方法名称").fill = header_fill
    ws.cell(row, 1).font = header_font
    ws.cell(row, 2, "平均惯性比率").fill = header_fill
    ws.cell(row, 2).font = header_font
    ws.cell(row, 3, "执行时间(秒)").fill = header_fill
    ws.cell(row, 3).font = header_font
    ws.cell(row, 4, "是否最佳").fill = header_fill
    ws.cell(row, 4).font = header_font
    
    row = 2
    for method, results_df in all_results.items():
        avg_ratio = results_df['inertia_ratio'].mean()
        exec_time = execution_time.get(method, 0)
        is_best = "是" if method == best_method else "否"
        
        ws.cell(row, 1, method)
        ws.cell(row, 2, f"{avg_ratio:.4f}")
        ws.cell(row, 3, f"{exec_time:.2f}")
        ws.cell(row, 4, is_best)
        row += 1
    
    # 调整列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    # 为每个方法创建详细工作表
    for method, results_df in all_results.items():
        ws = wb.create_sheet(method)
        
        # 写入表头
        headers = list(results_df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        # 写入数据
        for row_idx, row_data in enumerate(results_df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row_idx, col_idx, value)
        
        # 调整列宽
        for col in range(len(headers)):
            col_letter = chr(65 + col)  # A=65, B=66, etc.
            ws.column_dimensions[col_letter].width = 20
        
        # 添加本地 vs 全局对比总结
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList
        
        # 找到local_inertia和global_inertia的列位置
        if 'local_inertia' in headers and 'global_inertia' in headers:
            local_col = headers.index('local_inertia') + 1
            global_col = headers.index('global_inertia') + 1
            
            # 获取数据范围（排除Total行）
            data_start_row = 2
            data_end_row = len([r for r in results_df.values if r[0] != 'Total']) + 1
            
            # 创建对比图表
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = f"{method} - 本地惯性 vs 全局惯性对比"
            chart.y_axis.title = '惯性值'
            chart.x_axis.title = '站点'
            
            # 添加本地惯性数据
            local_data = Reference(ws, min_col=local_col, min_row=1, max_row=data_end_row, max_col=local_col)
            chart.add_data(local_data, titles_from_data=True)
            
            # 添加全局惯性数据
            global_data = Reference(ws, min_col=global_col, min_row=1, max_row=data_end_row, max_col=global_col)
            chart.add_data(global_data, titles_from_data=True)
            
            # 设置X轴标签（站点ID）
            sites = Reference(ws, min_col=1, min_row=2, max_row=data_end_row)
            chart.set_categories(sites)
            
            # 添加图表到工作表
            chart.width = 15
            chart.height = 10
            chart.anchor = 'A' + str(data_end_row + 3)
            ws.add_chart(chart)
    
    # 保存Excel文件
    excel_path = os.path.join(output_dir, 'all_results_summary.xlsx')
    wb.save(excel_path)
    logger.info(f"所有结果已整合到Excel文件: {excel_path}")
    
    # 生成可视化图表
    logger.info("\n### 生成可视化图表 ###\n")
    save_visualizations(all_results, output_dir)
    
    # 生成简化总结报告
    logger.info("\n### 生成总结报告 ###\n")
    summary_path = os.path.join(output_dir, 'summary_report.txt')
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("联邦聚类聚合总结报告\n")
        f.write("=" * 60 + "\n\n")
        f.write(f"最佳聚合方法: {best_method}\n")
        f.write(f"最佳平均惯性比率: {all_results[best_method]['inertia_ratio'].mean():.4f}\n")
        f.write(f"聚类数量: {n_clusters}\n")
        f.write(f"参与站点数: {len(sites)}\n\n")
        
        f.write("各方法性能对比:\n")
        f.write("-" * 40 + "\n")
        for method, results in all_results.items():
            f.write(f"{method}:\n")
            f.write(f"  平均惯性比率: {results['inertia_ratio'].mean():.4f}\n")
            f.write(f"  执行时间: {execution_time.get(method, 0):.2f}秒\n")
        
        f.write(f"\n总执行时间: {sum(execution_time.values()):.2f}秒\n")
    
    logger.info(f"总结报告已保存: {summary_path}")
    
    # 打印总结
    logger.info("\n" + "="*80)
    logger.info("Step 2 完成！")
    logger.info("="*80)
    logger.info(f"输出目录: {output_dir}")
    logger.info(f"总执行时间: {sum(execution_time.values()):.2f} 秒")
    logger.info("\n生成的文件:")
    logger.info(f"  - all_results_summary.xlsx: 所有结果的Excel汇总（包含方法对比和各方法详细数据）")
    logger.info(f"  - methods_comparison.png: 方法对比可视化图表")
    logger.info(f"  - [方法名]_详细分析.png: 各方法的详细分析图表")
    logger.info(f"  - best_global_centers.npy: 最佳聚合方法的全局聚类中心")
    logger.info(f"  - summary_report.txt: 文本总结报告")
    logger.info("="*80 + "\n")
    
    return sites, server, all_results, output_dir


if __name__ == "__main__":
    # 配置参数
    DATA_PATHS = {
        1: "E:/研三/py_ai_clinician-master/Dataset/dataave/group1_mimic_data.csv",
        2: "E:/研三/py_ai_clinician-master/Dataset/dataave/group2_mimic_data.csv",
        3: "E:/研三/py_ai_clinician-master/Dataset/dataave/group3_mimic_data.csv"
    }
    
    N_CLUSTERS = 750
    RANDOM_STATE = 42
    OUTPUT_DIR = "results"
    
    # 运行完整流程
    sites, server, results, output_dir = run_step2_complete_pipeline(
        data_paths=DATA_PATHS,
        n_clusters=N_CLUSTERS,
        random_state=RANDOM_STATE,
        output_base_dir=OUTPUT_DIR,
        use_first_bloc_only=True  # 只使用每个患者的第一条bloc数据
    )
    
    print("\n" + "="*80)
    print("Step 2 高级联邦聚类聚合完成！")
    print("="*80)
    print(f"\n查看结果目录: {output_dir}")
    print("\n主要输出文件:")
    print("  - all_results_summary.xlsx: 所有结果的Excel汇总（包含方法对比和各方法详细数据）")
    print("  - methods_comparison.png: 方法对比可视化图表")
    print("  - [方法名]_详细分析.png: 各方法的详细分析图表")
    print("  - best_global_centers.npy: 最佳聚合方法的全局聚类中心")
    print("  - summary_report.txt: 文本总结报告")
    print("="*80 + "\n")
